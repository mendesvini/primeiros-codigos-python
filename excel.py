import openpyxl


book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('frutas')

frutas_page = book['frutas']

frutas_page.append(['bananas','5','R$ 3,90'])
frutas_page.append(['bananas','5','R$ 3,90'])
frutas_page.append(['bananas','5','R$ 3,90'])
frutas_page.append(['bananas','5','R$ 3,90'])
frutas_page.append(['bananas','5','R$ 3,90'])
frutas_page.append(['bananas','5','R$ 3,90'])


book.save('Planilha de compras.xlsx')